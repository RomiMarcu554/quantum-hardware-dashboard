"""
write_to_db_web.py — Appends one paper's extracted parameters to the WEB
COPY of the platform Excel database (web_dashboard_version/data/*.xlsx).

IMPORTANT — SAFETY: This module only ever opens files inside this folder's
own `data/` directory (resolved relative to this file). It never touches
the original project files in `Final/` or `Anat/`.

Usage (also importable as a function from app.py):
    write_paper(data: dict) -> (success: bool, message: str)

`data` must look like:
{
    "platform": "Superconducting Circuits",
    "paper_title": "...",
    "authors": "...",
    "institution": "...",
    "year": 2017,
    "journal": "...",   (optional)
    "url": "...",       (optional)
    "parameters": {
        "T1": "7.6 us (8)",
        "Readout Fidelity": "98.42 +/- 0.07% (7)"
    }
}
"""

import openpyxl
from lib_data import FILE_MAP, PLATFORMS


def normalize(s):
    if s is None:
        return ""
    return str(s).replace("\n", " ").strip()


def write_paper(data: dict):
    platform = data.get("platform", "")
    if platform not in PLATFORMS:
        return False, f"Unknown platform. Must be one of: {PLATFORMS}"

    excel_path = FILE_MAP[platform]
    if not excel_path.exists():
        return False, f"Web copy of database not found: {excel_path}"

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers_raw = [cell.value for cell in ws[1]]
    headers = [normalize(h) for h in headers_raw]
    col_index = {h: i + 1 for i, h in enumerate(headers) if h}

    METADATA = {
        "Paper Title":                  data.get("paper_title", ""),
        "Authors":                      data.get("authors", ""),
        "Research Group / Institution": data.get("institution", ""),
        "Journal":                      data.get("journal", "") or "",
        "Year":                         data.get("year", ""),
        "URL":                          data.get("url", "") or "",
    }

    supplied_params = data.get("parameters", {}) or {}
    supplied_norm = {normalize(k): v for k, v in supplied_params.items()}

    new_row_idx = ws.max_row + 1
    filled, empty_list = [], []

    for header_norm, col_num in col_index.items():
        if not header_norm:
            continue
        if header_norm in METADATA:
            ws.cell(row=new_row_idx, column=col_num, value=METADATA[header_norm])
            continue
        if header_norm in supplied_norm and supplied_norm[header_norm]:
            cell_value = supplied_norm[header_norm]
            ws.cell(row=new_row_idx, column=col_num, value=cell_value)
            filled.append(f"{header_norm}: {cell_value}")
        else:
            empty_list.append(header_norm)

    wb.save(excel_path)

    msg_lines = [f"Row {new_row_idx} appended to web copy: {excel_path.name}",
                 f"Parameters written ({len(filled)}):"]
    msg_lines += [f"  • {line}" for line in filled]
    if empty_list:
        msg_lines.append(f"Left empty ({len(empty_list)}): {', '.join(empty_list)}")
    return True, "\n".join(msg_lines)
